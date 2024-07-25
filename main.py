import tkinter as tk
from fractions import Fraction
from tkinter import *
from tkinter import messagebox

import pandas as pd
from openpyxl import load_workbook

# Настройка pandas для правильного отображения
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', None)
pd.options.display.expand_frame_repr = False


def save_to_excel():  # сохраняет все в эксель и вызывает функции, преобразовывающие таблицу
    df = pd.DataFrame(columns=['А' + str(i + 1) for i in range(cols)])

    for i in range(rows):
        row_data = [int(entry_grid[i][j].get()) for j in range(cols)]
        df.loc[i] = row_data

    df.to_excel("data.xlsx", index=False)

    messagebox.showinfo("Save to Excel", "Data saved to data.xlsx")
    add_basis()
    add_CJ()
    shift_columns_right()
    add_numbers()
    add_Cbasis1stly()
    add_B1stly()
    add_Cj_forAB()
    rebuld_if_b_lowerzerro()
    delta_j1()
    delta_j2()
    i = 0
    readxl(i)
    create_boofer_sheet()
    count_while_ab(i)
    delete_basis()
    readxl2(i)
    count_while_cb(i)
    shift_columns_rightC2()
    add_C2()
    count_j2_for_C2()
    find_min_divj()


def create_grid():  # создает таблицу коэффициентов свободных переменных
    label = tk.Label(root, text="Введите ограничения равенства:")
    label.grid(row=2, column=0)
    for i in range(rows):
        row = []
        for j in range(cols):
            entry = tk.Entry(root, width=10)
            entry.grid(row=i + 3, column=j)
            row.append(entry)
        entry_grid.append(row)


root = tk.Tk()
root.title("Каноническая ЗЛП с параметром в целевой функции")
rows = 0
cols = 0
entry_grid = []
target_grid = []
target2_grid = []
B_grid = []
d_label_grid = []


def create_table():  # кнопка создать таблицу и поля ввода количества ограничений и переменных
    global rows, cols, entry_grid

    rows = int(rows_entry.get())
    cols = int(cols_entry.get())

    create_grid()

    save_button = tk.Button(root, text="Решить", command=save_to_excel)
    save_button.grid(row=rows + 3, columnspan=cols)
    create_C1()
    create_C2()
    create_B()


rows_label = tk.Label(root, text="Введите количество ограничений:")
rows_label.grid(row=0, column=0)
rows_entry = tk.Entry(root)
rows_entry.grid(row=0, column=1)

cols_label = tk.Label(root, text="Введите количество переменных:")
cols_label.grid(row=1, column=0)
cols_entry = tk.Entry(root)
cols_entry.grid(row=1, column=1)

create_button = tk.Button(root, text="Создать таблицу", command=create_table)
create_button.grid(row=1, column=2)


def add_basis():  # создает искуственный базис
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    print(rows, cols)
    for i in range(2, rows + 2):
        for j in range(cols + 1, cols + rows + 1):
            ws.cell(row=i, column=j).value = 0
            ws.cell(row=i, column=i + cols - 1).value = 1
    wb.save(fn)
    wb.close()


def shift_columns_right():  # перемещает столбцы вправо на 4 столбца, для добавления номеров таблиц,,переменных, добавленных в базис, коэфициентов базисных переменных и вектора B
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    sheet = wb['Sheet1']

    for row in sheet.iter_rows():
        for cell in row[::-1]:
            if cell.column > 0:
                new_cell = sheet.cell(row=cell.row, column=cell.column + 4, value=cell.value)
                cell.value = None

    wb.save(fn)
    wb.close()


def add_numbers():  # создает столбец с номерами
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws['A1'].value = '№'
    k = 1
    for i in range(2, rows + 2):
        ws.cell(row=i, column=1).value = k
        k += 1
    wb.save(fn)
    wb.close()


def add_CJ():  # добавляет СJ для колонок переменных целевой функции в таблице
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    cj = []
    for j in range(cols):
        cj.append(int(target_grid[j].get()))

    for j in range(1, cols + 1):
        ws.cell(row=1, column=j).value = str(ws.cell(row=1, column=j).value) + " " + str(cj[j - 1])
    wb.save(fn)
    wb.close()


def add_Cbasis1stly():  # Добавляет колонку С базиса в первый раз, (все -M т.к искусственный)
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws['B1'].value = 'Баз'
    ws['C1'].value = 'C` баз'
    for i in range(2, rows + 2):
        ws.cell(row=i, column=2).value = 'A' + str((cols + i - 1))
        ws.cell(row=i, column=3).value = '-M'
    wb.save(fn)
    wb.close()


def add_B1stly():  # Добавляет вектор B в колонку A0 когда мы создаем таблицу
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws['D1'].value = 'A0'
    B = []
    for i in range(rows):
        B.append(int(B_grid[i].get()))
    for i in range(2, rows + 2):
        ws.cell(row=i, column=4).value = int(B[i - 2])

    wb.save(fn)
    wb.close()


def add_Cj_forAB():  # Добавляет Сj в искуственный базис когда мы создаем таблицу
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    for j in range(cols + 5, cols + rows + 5):
        ws.cell(row=1, column=j).value = 'A' + str(j - 4) + ' -M'
    wb.save(fn)
    wb.close()


def rebuld_if_b_lowerzerro():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    for i in range(2, rows + 2):
        if (ws['D' + str(i)].value < 0):
            for j in range(4, cols + 5):
                ws.cell(row=i, column=j).value = ws.cell(row=i, column=j).value * (-1)

    wb.save(fn)
    wb.close()


# создают поля ввода векторов C` C`` B
def create_C1():
    target_label = tk.Label(root, text="Введите С`:")
    target_label.grid(row=rows + 4, column=0)
    create_target()


def create_C2():
    target2_label = tk.Label(root, text="Введите С``:")
    target2_label.grid(row=rows + 5, column=0)
    create_target2()


def create_B():
    create_b()


def create_target():
    for j in range(cols):
        target = tk.Entry(root, width=5)
        target.grid(row=rows + 4, column=j + 1)
        target_grid.append(target)


def create_target2():
    for j in range(cols):
        target2 = tk.Entry(root, width=5)
        target2.grid(row=rows + 5, column=j + 1)
        target2_grid.append(target2)


def create_b():
    for i in range(rows):
        B = tk.Entry(root, width=5)
        B.grid(row=i + 3, column=cols + 2)
        B_lable = tk.Label(root, text="=           ")
        B_lable.grid(row=i + 3, column=cols + 1, padx=0)
        B_grid.append(B)


def delta_j2():  # считает делтьту j 2 черты
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws['C' + str(rows + 2)].value = 'Δj`'
    ws['C' + str(rows + 3)].value = 'Δj``'
    anulspot = Fraction('0/1')
    for i in range(2, rows + 2):
        if (ws['C' + str(i)].value == '-M'):
            sklr = Fraction(ws.cell(row=i, column=4).value)
            sklr = (sklr) * -1
            print(sklr)
            anulspot += sklr
            ws['D' + str(rows + 3)].value = str(anulspot)
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    for j in range(5, cols + 5):
        for i in range(2, rows + 2):
            if (ws['C' + str(i)].value == '-M'):
                sklr = Fraction(ws.cell(row=i, column=j).value)
                sklr = (sklr) * -1
                anulspot += sklr
                ws.cell(row=rows + 3, column=j).value = str(anulspot)
        sklr = Fraction('0/1')
        anulspot = Fraction('0/1')
    wb.save(fn)
    wb.close()


def find_target_element_when_ab():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    k = 0
    min = Fraction(str(ws['E' + str(rows + 3)].value))
    for j in range(5, cols + 5):
        if Fraction(ws.cell(row=rows + 3, column=j).value) <= min:
            min = Fraction(ws.cell(row=rows + 3, column=j).value)
            k = j
    print(min, k)

    mindiv = Fraction('0/1')
    for i in range(2, rows + 5):
        if (ws['C' + str(i)].value == '-M'):
            if (Fraction(ws.cell(row=i, column=k).value) != 0):
                if (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value) > 0):
                    mindiv = Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value)
                    print(mindiv)
                    break
    targetel = Fraction('0/1')

    g = 0

    for i in range(2, rows + 3):
        if (ws['C' + str(i)].value == '-M'):
            if (Fraction(ws.cell(row=i, column=k).value) != 0):
                if (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value) > 0):
                    if ((Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value)) <= mindiv):
                        targetel = Fraction(ws.cell(row=i, column=k).value)
                        mindiv = (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value))
                        g = i
    print(targetel, g)

    wb.save(fn)
    wb.close()
    new_table(k, g, targetel)


def new_table(k, g, targetel):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    wss = wb['Sheet2']
    v = 0

    for i in range(2, rows + 2):
        for j in range(4, rows + cols + 5):
            if (i != g):
                A = Fraction(str(ws.cell(row=i, column=j).value))
                B = Fraction(str(ws.cell(row=i, column=k).value))
                C = Fraction(str(ws.cell(row=g, column=j).value))
                D = Fraction(str(targetel))
                print(v, ' ', A, '-', B, '*', C, '/', D)
                R = (A - ((B * C) / D))
                wss.cell(row=i, column=j).value = str(R)
                print(R)
                v += 1
    for i in range(2, rows + 2):
        for j in range(4, rows + cols + 5):
            if (i != g):
                ws.cell(row=i, column=j).value = wss.cell(row=i, column=j).value
    for i in range(2, rows + 2):
        for j in range(4, rows + cols + 5):
            if (i == g):
                ws.cell(row=i, column=j).value = str(
                    Fraction(str(ws.cell(row=i, column=j).value)) / Fraction(str(targetel)))
                print(v, ' ', Fraction(str(ws.cell(row=i, column=j).value)), '/', Fraction(str(targetel)), ' ', '=',
                      ' ', Fraction(str(ws.cell(row=i, column=j).value)))
                v += 1
    wb.save(fn)
    wb.close()
    replace_basis(k, g)


def replace_basis(k, g):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws.cell(row=g, column=2).value = 'A' + str(k - 4)
    cj = []
    for j in range(cols):
        cj.append(int(target_grid[j].get()))

    ws.cell(row=g, column=3).value = cj[k - 5]
    wb.save(fn)
    wb.close()


def readxl(i):
    dt = pd.read_excel("data.xlsx", sheet_name="Sheet1")
    d_text = Text(root, height=7, width=40, wrap=NONE)
    d_text.insert(END, str(dt))
    d_text.grid(row=rows + 7 + i, column=0)
    hsb = Scrollbar(root, orient="horizontal", command=d_text.xview)
    d_text.configure(xscrollcommand=hsb.set)
    hsb.grid(row=rows + 8 + i, column=0, sticky="ew")
    print(dt)


def readxl2(i):
    dt = pd.read_excel("data.xlsx", sheet_name="Sheet1")
    d_text = Text(root, height=7, width=40, wrap=NONE)
    d_text.insert(END, str(dt))
    d_text.grid(row=rows + 7 + i, column=1)
    hsb = Scrollbar(root, orient="horizontal", command=d_text.xview)
    d_text.configure(xscrollcommand=hsb.set)
    hsb.grid(row=rows + 8 + i, column=1, sticky="ew")
    print(dt)


def readxl3(i):
    dt = pd.read_excel("data.xlsx", sheet_name="Sheet1")
    d_text = Text(root, height=7, width=40, wrap=NONE)
    d_text.insert(END, str(dt))
    d_text.grid(row=rows + 7 + i, column=2)
    hsb = Scrollbar(root, orient="horizontal", command=d_text.xview)
    d_text.configure(xscrollcommand=hsb.set)
    hsb.grid(row=rows + 8 + i, column=2, sticky="ew")
    print(dt)


def readxl4(i, lv, ln):
    d_text = tk.Text(root, height=2, width=20, padx=1)
    x = '['
    y = ']'
    if (lv) == '+∞':
        y = ')'
    if (ln) == '-∞':
        x = '('
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    z = Fraction(ws['E' + str(rows + 2)].value)
    w = Fraction(ws['E' + str(rows + 3)].value)
    if Fraction(w) > 0:
        w = ('+' + str(w))
    wb.save(fn)
    wb.close()
    d_text.insert(tk.END, ('λ ∈ ' + str(x) + str(ln) + ';' + str(lv) + str(y)))
    d_text.grid(row=rows + 16 + i, column=0)

    e_text = tk.Text(root, height=2, width=20, padx=1)
    e_text.insert(tk.END, ('l(λ)=' + str(z) + str(w) + 'λ'))
    e_text.grid(row=rows + 17 + i, column=0)


def readxl5(lv, ln):
    d_text = tk.Text(root, height=2, width=20, padx=1)
    x = '['
    y = ']'
    if lv == '+∞':
        y = ')'
    if ln == '-∞':
        x = '('
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    z = Fraction(ws['E' + str(rows + 2)].value)
    w = Fraction(ws['E' + str(rows + 3)].value)
    if (Fraction(w) > 0):
        w = ('+' + str(w))
    wb.save(fn)
    wb.close()
    d_text.insert(tk.END, ('λ ∈ ' + str(x) + str(ln) + ';' + str(lv) + str(y)))
    d_text.grid(row=rows + 16, column=1)
    e_text = tk.Text(root, height=2, width=20, padx=1)
    e_text.insert(tk.END, ('l(λ)=' + str(z) + str(w) + 'λ'))
    e_text.grid(row=rows + 17, column=1)


def readlv(lv):
    d_text = tk.Text(root, height=2, width=20, padx=1)
    d_text.insert(tk.END, ('[ ' + str(lv) + '; +∞ )'))
    d_text.grid(row=rows + 16, column=2)


def readln(ln):
    d_text = tk.Text(root, height=2, width=20, padx=1)
    d_text.insert(tk.END, ('(-∞; ' + str(ln) + ' ]'))
    d_text.grid(row=rows + 17, column=2)


def if_ab():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    for i in range(2, rows + 2):
        if ((ws['C' + str(i)].value) == '-M'):
            return False
    else:
        return True


def if_lower_than_low():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    for j in range(5, cols + 4):
        if Fraction(ws.cell(row=rows + 2, column=j).value) < Fraction('0/1'):
            return False
    else:
        return True


def count_while_ab(i):
    x = if_ab()
    if (x == False):
        find_target_element_when_ab()
        delta_j1()
        delta_j2()
        i += 2
        readxl(i)
        count_while_ab(i)
    else:
        return 0


def count_while_cb(i):
    x = if_lower_than_low()
    if (x == False):
        find_targetel_when_cb()
        delta_j1()
        i += 2
        readxl2(i)
        count_while_cb(i)
    else:
        return 0


def create_boofer_sheet():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb.create_sheet(title='Sheet2')
    wb.save(fn)
    wb.close()


def delete_basis():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    for i in range(1, rows + 3):
        for j in range(5 + cols, cols + rows + 5):
            ws.cell(row=i, column=j).value = None
    for j in range(1, cols + rows + 5):
        ws.cell(row=(rows + 3), column=j).value = None

    wb.save(fn)
    wb.close()


def delta_j1():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    for i in range(2, rows + 2):
        if (ws['C' + str(i)].value != '-M'):
            sklr = Fraction(ws.cell(row=i, column=4).value)
            sklr = (sklr) * Fraction(ws['C' + str(i)].value)
            anulspot += sklr
        else:
            sklr = Fraction('0/1')
            anulspot += sklr
    ws['D' + str(rows + 2)].value = str(anulspot)
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    cj = []
    for j in range(cols):
        cj.append(int(target_grid[j].get()))
    for j in range(5, cols + 5):
        for i in range(2, rows + 2):
            if (ws['C' + str(i)].value != '-M'):
                sklr = Fraction(ws.cell(row=i, column=j).value)
                sklr = (sklr) * Fraction(ws['C' + str(i)].value)
                anulspot += sklr
            else:
                sklr = Fraction('0/1')
                anulspot = sklr
        ws.cell(row=rows + 2, column=j).value = str(anulspot - Fraction(str(cj[j - 5])))
        sklr = Fraction('0/1')
        anulspot = Fraction('0/1')
    wb.save(fn)
    wb.close()


def find_targetel_when_cb():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    k = 0
    min = Fraction(ws['E' + str(rows + 2)].value)
    for j in range(5, cols + 5):
        if Fraction(ws.cell(row=rows + 2, column=j).value) <= min:
            min = Fraction(ws.cell(row=rows + 2, column=j).value)
            k = j
    print(min, k)

    mindiv = Fraction('0/1')
    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=k).value) != 0):
            if (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value) > 0):
                mindiv = Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value)
    targetel = Fraction('0/1')

    g = 0

    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=k).value) != 0):
            if (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value) <= mindiv):
                if (Fraction(ws['D' + str(i)].value) / Fraction(ws.cell(row=i, column=k).value) > 0):
                    targetel = Fraction(ws.cell(row=i, column=k).value)
                    g = i
    z = 0
    for j in range(5, cols + 5):
        if (Fraction(ws.cell(row=rows + 2, column=j).value) >= 0):
            z += 1
    if (z == cols):
        shift_columns_rightC2()
        add_C2()
        count_j2_for_C2()
        find_min_divj()
        return
    print(targetel, g)
    wb.save(fn)
    wb.close()
    new_table2(k, g, targetel)


def new_table2(k, g, targetel):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    wss = wb['Sheet2']
    v = 0

    for i in range(2, rows + 2):
        for j in range(4, cols + 4):
            if (i != g):
                A = Fraction(str(ws.cell(row=i, column=j).value))
                B = Fraction(str(ws.cell(row=i, column=k).value))
                C = Fraction(str(ws.cell(row=g, column=j).value))
                D = Fraction(str(targetel))
                print(v, ' ', A, '-', B, '*', C, '/', D)
                R = (A - ((B * C) / D))
                wss.cell(row=i, column=j).value = str(R)
                print(R)
                v += 1
    for i in range(2, rows + 2):
        for j in range(4, cols + 4):
            if (i != g):
                ws.cell(row=i, column=j).value = wss.cell(row=i, column=j).value
    for i in range(2, rows + 2):
        for j in range(4, cols + 4):
            if (i == g):
                ws.cell(row=i, column=j).value = str(
                    Fraction(str(ws.cell(row=i, column=j).value)) / Fraction(str(targetel)))
                print(v, ' ', Fraction(str(ws.cell(row=i, column=j).value)), '/', Fraction(str(targetel)), ' ', '=',
                      ' ', Fraction(str(ws.cell(row=i, column=j).value)))
                v += 1
    wb.save(fn)
    wb.close()
    replace_basis(k, g)


def shift_columns_rightC2():  # перемещает столбцы вправо на 1 столбец, для добавления С2
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    sheet = wb['Sheet1']

    for row in sheet.iter_rows():
        for cell in row[::-1]:
            if cell.column > 3:
                new_cell = sheet.cell(row=cell.row, column=cell.column + 1, value=cell.value)
                cell.value = None
    wb.save(fn)
    wb.close()


def add_C2():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    cj2 = []
    for j in range(cols):
        cj2.append(int(target2_grid[j].get()))
    for i in range(2, rows + 2):
        j = list(str(ws['B' + str(i)].value))
        ws['D' + str(i)].value = cj2[int(j[1]) - 1]
    ws['D1'].value = 'C``баз'
    ws['D' + str(rows + 3)].value = 'Δj``'
    for j in range(0, cols):
        ws.cell(row=1, column=6 + j).value = str(ws.cell(row=1, column=6 + j).value) + '|' + str(cj2[j])
    wb.save(fn)
    wb.close()


def count_j2_for_C2():
    delta_j1_in_the_end()
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']

    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    for i in range(2, rows + 2):
        sklr = Fraction(ws.cell(row=i, column=5).value)
        sklr = (sklr) * Fraction(ws['D' + str(i)].value)
        anulspot += sklr
    ws['E' + str(rows + 3)].value = str(anulspot)
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    cj2 = []
    for j in range(cols):
        cj2.append(int(target2_grid[j].get()))
    for j in range(6, cols + 6):
        for i in range(2, rows + 2):
            sklr = Fraction(ws.cell(row=i, column=j).value)
            sklr = (sklr) * Fraction(ws['D' + str(i)].value)
            anulspot += sklr
        ws.cell(row=rows + 3, column=j).value = str(anulspot - Fraction(str(cj2[j - 6])))
        sklr = Fraction('0/1')
        anulspot = Fraction('0/1')
    ws['D' + str(rows + 2)].value = None
    wb.save(fn)
    wb.close()


def check_a_diffrence_j2():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    plus = 0
    minus = 0
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) >= 0):
            plus += 1
        if (Fraction(ws.cell(row=rows + 3, column=j).value) <= 0):
            minus += 1
    if (plus == cols):
        return 1
    if (minus == cols):
        print('mines---')
        return 0
    else:
        return 2


def find_min_divj():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    minj = Fraction('0/1')
    maxj = Fraction('0/1')

    lv = Fraction('0/1')
    ln = Fraction('0/1')
    nv = 0
    nn = 0
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) <= lv):
                    lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nv = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) >= 0):
            if (check_a_diffrence_j2() == 1):
                lv = str('+∞')
    print('lv ', lv, nv)

    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) >= ln):
                    ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nn = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) <= 0):
            if (check_a_diffrence_j2() == 0):
                ln = str('-∞')
    print('ln ', ln, nn)
    wb.save(fn)
    wb.close()
    check_all(nn, ln, nv, lv)


def check_all(nn, ln, nv, lv):
    i = 0
    readxl3(i)
    readxl4(i, lv, ln)
    if (check_4_positive_n(nn, ln) == True):
        if (check_4_positive_v(nv, lv) == False):
            print('Нет решений при  λ>' + str(lv))
            readlv(lv)
            try:
                mindivwhenl(nn)
                find_min_divj_4_n()

            except:
                return

    if (check_4_positive_n(nn, ln) == False):
        if (check_4_positive_v(nv, lv) == True):
            print('Нет решений при  λ<' + str(ln))
            readln(ln)

            try:
                mindivwhenl(nv)
                find_min_divj_4_v()

            except:
                return


    else:
        i += 1
        print('Нет решений при  λ>' + str(lv))
        print('Нет решений при  λ<' + str(ln))

        return


def mindivwhenl(x):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    mindiv = Fraction('0/1')
    y = 0
    targetel = Fraction('0/1')

    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=x).value) != 0):
            if (Fraction(ws['E' + str(i)].value) / Fraction(ws.cell(row=i, column=x).value) > 0):
                mindiv = Fraction(ws['E' + str(i)].value) / Fraction(ws.cell(row=i, column=x).value)
                break
    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=x).value) != 0):
            if (Fraction(ws['E' + str(i)].value) / Fraction(ws.cell(row=i, column=x).value) > 0):
                if (Fraction(ws['E' + str(i)].value) / Fraction(ws.cell(row=i, column=x).value) <= mindiv):
                    mindiv = Fraction(ws['E' + str(i)].value) / Fraction(ws.cell(row=i, column=x).value)
                    y = i

    targetel = Fraction(ws.cell(row=y, column=x).value)
    print(targetel, ' ', x, ' ', y)
    new_tableee(x, y, targetel)
    i = 2
    readxl3(i)


def check_4_positive_n(nn, ln):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    k = 0
    if ln == '-∞':
        return False
    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=nn).value) < 0):
            k += 1
    if (k == rows):
        return False
    else:
        return True


def check_4_positive_v(nv, lv):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    k = 0
    if lv == '+∞':
        return False
    for i in range(2, rows + 2):
        if (Fraction(ws.cell(row=i, column=nv).value) < 0):
            k += 1
    if (k == rows):
        return False
    else:
        return True


def reducebas(x, y):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    ws.cell(row=y, column=2).value = 'A' + str(x - 5)
    cj = []
    cj2 = []
    for j in range(cols):
        cj.append(int(target_grid[j].get()))
    for j in range(cols):
        cj2.append(int(target2_grid[j].get()))

    ws.cell(row=y, column=3).value = cj[x - 6]
    ws.cell(row=y, column=4).value = cj2[x - 6]
    wb.save(fn)
    wb.close()


def new_tableee(x, z, targetel):
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    wss = wb['Sheet2']
    v = 0
    for i in range(2, rows + 2):
        for j in range(5, cols + 6):
            if (i != x):
                A = Fraction(str(ws.cell(row=i, column=j).value))
                B = Fraction(str(ws.cell(row=i, column=x).value))
                C = Fraction(str(ws.cell(row=z, column=j).value))
                D = Fraction(str(targetel))
                print(v, ' ', A, '-', B, '*', C, '/', D)
                R = (A - ((B * C) / D))
                wss.cell(row=i, column=j).value = str(R)
                print(R)
                v += 1
    for i in range(2, rows + 2):
        for j in range(5, cols + 6):
            if (i != z):
                ws.cell(row=i, column=j).value = wss.cell(row=i, column=j).value
    for i in range(2, rows + 2):
        for j in range(5, cols + 6):
            if (i == z):
                ws.cell(row=i, column=j).value = str(
                    Fraction(str(ws.cell(row=i, column=j).value)) / Fraction(str(targetel)))
                print(v, ' ', Fraction(str(ws.cell(row=i, column=j).value)), '/', Fraction(str(targetel)), ' ', '=',
                      ' ', Fraction(str(ws.cell(row=i, column=j).value)))
                v += 1
    wb.save(fn)
    wb.close()
    reducebas(x, z)
    count_j2_for_C2()


def find_min_divj_4_v():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    minj = Fraction('0/1')
    maxj = Fraction('0/1')

    lv = Fraction('0/1')
    ln = Fraction('0/1')
    nv = 0
    nn = 0
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) <= lv):
                    lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nv = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) >= 0):
            if (check_a_diffrence_j2() == 1):
                lv = str('+∞')
    print('lv ', lv, nv)

    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) >= ln):
                    ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nn = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) <= 0):
            if (check_a_diffrence_j2() == 0):
                ln = str('-∞')
    print('ln ', ln, nn)
    wb.save(fn)
    wb.close()
    if (check_4_positive_v(nv, lv) == True):
        mindivwhenl(nv)
    else:
        print('Нет решений при  λ>' + str(lv))
        readxl5(lv, ln)
        readlv(lv)
        return


def find_min_divj_4_n():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    minj = Fraction('0/1')
    maxj = Fraction('0/1')

    lv = Fraction('0/1')
    ln = Fraction('0/1')
    nv = 0
    nn = 0
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) < 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) <= lv):
                    lv = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nv = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) >= 0):
            if (check_a_diffrence_j2() == 1):
                lv = str('+∞')
    print('lv ', lv, nv)

    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(ws.cell(row=rows + 3, column=j).value)
                break
    for j in range(6, cols + 6):
        if (Fraction(ws.cell(row=rows + 3, column=j).value) != 0):
            if (Fraction(ws.cell(row=rows + 3, column=j).value) > 0):
                if ((-Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)) >= ln):
                    ln = -Fraction(ws.cell(row=rows + 2, column=j).value) / Fraction(
                        ws.cell(row=rows + 3, column=j).value)
                    nn = j
        if (Fraction(ws.cell(row=rows + 3, column=j).value) <= 0):
            if (check_a_diffrence_j2() == 0):
                ln = str('-∞')
    print('ln ', ln, nn)
    wb.save(fn)
    wb.close()
    if (check_4_positive_n(nn, ln) == True):
        mindivwhenl(nv)
    else:
        print('Нет решений при  λ<' + str(ln))
        readxl5(lv, ln)
        readln(ln)
        return


def delta_j1_in_the_end():
    fn = 'data.xlsx'
    wb = load_workbook(fn)
    ws = wb['Sheet1']
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    for i in range(2, rows + 2):
        sklr = Fraction(ws.cell(row=i, column=5).value)
        sklr = (sklr) * Fraction(ws['C' + str(i)].value)
        anulspot += sklr

    ws['E' + str(rows + 2)].value = str(anulspot)
    sklr = Fraction('0/1')
    anulspot = Fraction('0/1')
    cj = []
    for j in range(cols):
        cj.append(int(target_grid[j].get()))
    for j in range(6, cols + 6):
        for i in range(2, rows + 2):
            sklr = Fraction(ws.cell(row=i, column=j).value)
            sklr = (sklr) * Fraction(ws['C' + str(i)].value)
            anulspot += sklr
        ws.cell(row=rows + 2, column=j).value = str(anulspot - Fraction(str(cj[j - 6])))
        sklr = Fraction('0/1')
        anulspot = Fraction('0/1')
    wb.save(fn)
    wb.close()


root.mainloop()
